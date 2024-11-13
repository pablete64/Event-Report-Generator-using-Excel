import tkinter as tk
from tkinter import filedialog
from tkinter import simpledialog
import pandas as pd
import os
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ignorar advertencias específicas sobre expresiones regulares
warnings.filterwarnings("ignore", message="This pattern has match groups")

def seleccionar_archivo():
    root = tk.Tk()
    root.withdraw()  # Ocultamos la ventana principal de Tkinter
    file_path = filedialog.askopenfilename()
    return file_path

def seleccionar_rango_fechas():
    root = tk.Tk()
    root.withdraw()  # Ocultamos la ventana principal de Tkinter
    fecha_inicial = simpledialog.askstring("Fecha inicial", "Introduce la fecha inicial (DD-MM-YYYY):", parent=root)
    fecha_final = simpledialog.askstring("Fecha final", "Introduce la fecha final (DD-MM-YYYY):", parent=root)
    fecha_inicial = pd.to_datetime(fecha_inicial, format='%d-%m-%Y')
    fecha_final = pd.to_datetime(fecha_final, format='%d-%m-%Y')
    return fecha_inicial, fecha_final

def calcular_dias_disparo_por_entrada(df, fecha_inicial, fecha_final):
    df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True)
    df.sort_values(by=['Fecha', 'Hora'], ascending=True, inplace=True)

    dias_con_disparo = {}

    for entry_num in range(1, 17):
        total_dias = 0
        fecha_analizada = pd.to_datetime('1900-01-01')  # Fecha ficticia para iniciar el análisis

        entry_df = df[df['DESCRIPCION'].str.contains(f'Ent\\({entry_num}\\)')].copy()

        # Verificar si el primer evento es un "Res" y ajustar el conteo inicial de días
        if not entry_df.empty:
            primer_evento = entry_df.iloc[0]
            if 'Res' in primer_evento['DESCRIPCION'] and primer_evento['Fecha'] > fecha_inicial:
                # Calcular días desde la fecha inicial hasta el primer "Res" incluyendo el día del "Res"
                total_dias += (primer_evento['Fecha'] - fecha_inicial).days + 1
                fecha_analizada = primer_evento['Fecha']

        for fecha, grupo in entry_df.groupby(entry_df['Fecha']):
            if fecha <= fecha_analizada:
                continue  # Ignorar fechas ya analizadas
            dis_del_dia = grupo[grupo['DESCRIPCION'].str.contains('Dis')]
            if not dis_del_dia.empty:
                fecha_dis = dis_del_dia.iloc[0]['Fecha']
                ultimo_evento_del_dia = grupo.iloc[-1]
                fecha_ultimo_evento = ultimo_evento_del_dia['Fecha']

                if 'Res' in ultimo_evento_del_dia['DESCRIPCION']:
                    fecha_analizada = fecha_ultimo_evento
                    total_dias += 1  # Concluir ese día como 1 día de disparo
                else:
                    res_subsiguiente = entry_df[(entry_df['Fecha'] > fecha_ultimo_evento) & (entry_df['DESCRIPCION'].str.contains('Res'))].head(1)
                    if not res_subsiguiente.empty:
                        fecha_res = res_subsiguiente.iloc[0]['Fecha']
                        total_dias += (fecha_res - fecha_dis).days + 1
                        fecha_analizada = fecha_res
                    else:
                        total_dias += 1  # Contar como un día si no hay "Res" subsiguiente
                        fecha_analizada = fecha_dis  # Actualizar la fecha analizada para continuar

        # Comprobar si el último evento de la entrada es un "Dis" y ajustar el total de días con disparo
        if not entry_df.empty:
            ultimo_evento = entry_df.iloc[-1]
            if 'Dis' in ultimo_evento['DESCRIPCION']:
                diferencia = (fecha_final - ultimo_evento['Fecha']).days
                total_dias += diferencia + 1

        dias_con_disparo[f'Ent({entry_num})'] = total_dias

    return dias_con_disparo


def analizar_eventos(file_path, fecha_inicial, fecha_final):
    df = pd.read_excel(file_path)

    # Filtrar por el rango de fechas
    df['FechaHora'] = pd.to_datetime(df['Fecha'] + ' ' + df['Hora'], dayfirst=True)
    df = df[(df['FechaHora'] >= fecha_inicial) & (df['FechaHora'] < fecha_final + pd.Timedelta(days=1))]

    eventos_filtrados = df.copy()
    eventos_filtrados.sort_values(by='FechaHora', inplace=True)

    # Preparar datos para el resumen
    resumen_data = {
        'Entrada': [],
        'Nº Disparos': [],
        'Res > 1H': [],
        'Res > 1 Día': [],
        'Nº de días con Disparo': [],
    }

    for entry_num in range(1, 17):
        total_dias = calcular_dias_disparo_por_entrada(eventos_filtrados, fecha_inicial, fecha_final)
        resumen_data['Nº de días con Disparo'].append(total_dias[f'Ent({entry_num})'])

    def emparejar_eventos_y_calcular_excesos(entrada_num):
        eventos = eventos_filtrados[eventos_filtrados['DESCRIPCION'].str.contains(f'Ent\\({entrada_num}\\)')].copy()
        disparos = eventos[eventos['DESCRIPCION'].str.contains('Dis')].copy()
        rearmes = eventos[eventos['DESCRIPCION'].str.contains('Res')].copy()

        parejas = []
        num_disparos = 0  # Inicializar contador de disparos

        for _, fila in disparos.iterrows():
            num_disparos += fila['DESCRIPCION'].count('Dis')  # Contar cada "Dis" en la descripción

        for _, disparo in disparos.iterrows():
            rearme_cercano = rearmes[rearmes['FechaHora'] >= disparo['FechaHora']].head(1)
            if not rearme_cercano.empty:
                rearme = rearme_cercano.iloc[0]
                diferencia = rearme['FechaHora'] - disparo['FechaHora']
                excede_limite_horas = diferencia.total_seconds() > 3600
                excede_dia_siguiente = (rearme['FechaHora'].date() > disparo['FechaHora'].date())
                parejas.append([disparo['Fecha'], disparo['Hora'], rearme['Fecha'], rearme['Hora'], excede_limite_horas, excede_dia_siguiente])

        res_mas_24h = sum([par[5] for par in parejas])
        res_mas_1h = sum([par[4] for par in parejas])

        resumen_data['Entrada'].append(f'Ent({entrada_num})')
        resumen_data['Nº Disparos'].append(num_disparos)
        resumen_data['Res > 1H'].append(res_mas_1h)
        resumen_data['Res > 1 Día'].append(res_mas_24h)

        return pd.DataFrame(parejas, columns=['Fecha Disparo', 'Hora Disparo', 'Fecha Rearme', 'Hora Rearme', 'Excede 1 hora', 'Excede al día siguiente'])

    # Nombre del archivo de salida
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_file_path = os.path.join(os.path.dirname(file_path), f'{base_name}_revisado.xlsx')

    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        workbook = writer.book
        red_format = workbook.add_format({'bg_color': '#FFC7CE'})
        yellow_format = workbook.add_format({'bg_color': '#FFFF00'})
        grey_format = workbook.add_format({'bg_color': '#D3D3D3'})

        # Obtener el nombre base del archivo
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Exportar eventos filtrados sin la columna 'FechaHora' y cambiar el nombre de la hoja
        eventos_filtrados.drop(columns=['FechaHora']).to_excel(writer, sheet_name=base_name, index=False)

        for i in range(1, 17):
            # Procesar y exportar resumen para cada entrada
            df_entrada_resumen = emparejar_eventos_y_calcular_excesos(i)
            if not df_entrada_resumen.empty:
                df_entrada_resumen.to_excel(writer, sheet_name=f'ENT({i}) Resumen', index=False)
                worksheet_resumen = writer.sheets[f'ENT({i}) Resumen']
                worksheet_resumen.conditional_format('E2:E{}'.format(len(df_entrada_resumen) + 1), {'type': 'cell', 'criteria': '==', 'value': 'TRUE', 'format': red_format})
                worksheet_resumen.conditional_format('F2:F{}'.format(len(df_entrada_resumen) + 1), {'type': 'cell', 'criteria': '==', 'value': 'TRUE', 'format': yellow_format})

            # Exportar datos filtrados por entrada y aplicar formato gris a las filas con "Ent(x) Dis"
            df_entrada_detalle = df[df['DESCRIPCION'].str.contains(f'Ent\\({i}\\)')].copy()
            if not df_entrada_detalle.empty:
                sheet_name_detalle = f'ENT({i}) Detalle'
                df_entrada_detalle.to_excel(writer, sheet_name=sheet_name_detalle, index=False)
                worksheet_detalle = writer.sheets[sheet_name_detalle]
                for row, desc in enumerate(df_entrada_detalle['DESCRIPCION'], start=2):
                    if f"Ent({i}) Dis" in desc:
                        worksheet_detalle.set_row(row-1, None, grey_format)

        # Convertir datos de resumen a DataFrame y exportarlo como la primera hoja
        resumen_df = pd.DataFrame(resumen_data)
        resumen_df.to_excel(writer, sheet_name='RESUMEN', index=False)  # No es necesario ajuste de coordenadas

        # Ajustar ancho de columnas en todas las hojas
        for sheet in writer.sheets.values():
            for col in range(df.shape[1]):
                max_len = max(df.iloc[:, col].astype(str).map(len).max(), len(df.columns[col])) + 1
                sheet.set_column(col, col, max_len * 1.5)  # Aumentar el ancho multiplicando por un factor

    def reordenar_hojas_excel(path_archivo):
        workbook = load_workbook(path_archivo)
        sheet_names = workbook.sheetnames
        resumen_index = sheet_names.index('RESUMEN')
        sheet_names.insert(0, sheet_names.pop(resumen_index))  # Quita 'RESUMEN' y lo inserta al inicio
        workbook._sheets = [workbook[sheet] for sheet in sheet_names]
        workbook.save(path_archivo)
        resumen_sheets = [sheet for sheet in sheet_names if "Resumen" in sheet and sheet.startswith("ENT(")]
        non_resumen_sheets = [workbook[sheet] for sheet in sheet_names if sheet not in resumen_sheets]
        resumen_sheets_objs = [workbook[sheet] for sheet in resumen_sheets]

        # Reordenar hojas poniendo las de resumen al final
        workbook._sheets = non_resumen_sheets + resumen_sheets_objs
        workbook.save(path_archivo)

    reordenar_hojas_excel(output_file_path)

    return output_file_path

def main():
    archivo = seleccionar_archivo()
    if archivo:
        fecha_inicial, fecha_final = seleccionar_rango_fechas()
        resultados_path = analizar_eventos(archivo, fecha_inicial, fecha_final)
        
        # Modificar nombre de la primera hoja del Excel generado
        workbook = load_workbook(resultados_path)
        sheet_names = workbook.sheetnames
        entry_names = {}  # Diccionario para almacenar los nombres correspondientes a cada entrada

        for sheet_name in sheet_names:
            if sheet_name.startswith('ENT('):
                entry_number = sheet_name.split('(')[1].split(')')[0]
                df = pd.read_excel(resultados_path, sheet_name=sheet_name)
                if not df.empty and 'DESCRIPCION' in df.columns:
                    discharge_description = df[df['DESCRIPCION'].str.contains(f'Ent\\({entry_number}\\) Dis')]
                    if not discharge_description.empty:
                        first_discharge_description = discharge_description['DESCRIPCION'].iloc[0]
                        if len(first_discharge_description) >= 15:
                            index_dis = first_discharge_description.find('Dis:')
                            entry_name = f'Ent({entry_number}) {first_discharge_description[index_dis+4:index_dis+15]}'
                        else:
                            entry_name = f'Ent({entry_number}) {first_discharge_description}'
                        entry_names[entry_number] = entry_name
                    else:
                        entry_names[entry_number] = f'Ent({entry_number})'
        
        # Modificar nombres en la tabla RESUMEN y aplicar formato gris según corresponda
        resumen_sheet = workbook['RESUMEN']
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        for i, entry_number in enumerate(range(1, 17), start=2):  # Comenzar desde la fila 2
            if str(entry_number) in entry_names:
                resumen_sheet.cell(row=i, column=1, value=entry_names[str(entry_number)])  # Modificar la columna 1
                # Aplicamos el formato gris a toda la fila donde se ha asignado un alias
                for col in range(1, resumen_sheet.max_column + 1):
                    resumen_sheet.cell(row=i, column=col).fill = grey_fill
        
        # Guardar el archivo antes de ajustar el ancho de las columnas
        workbook.save(resultados_path)
        
        # Aumentar el ancho de las columnas en la hoja "RESUMEN"
        worksheet_resumen = workbook['RESUMEN']
        for col in range(1, worksheet_resumen.max_column + 1):
            max_len = max(len(str(cell.value)) for column in worksheet_resumen.iter_cols(min_col=col, max_col=col) for cell in column)
            worksheet_resumen.column_dimensions[worksheet_resumen.cell(row=1, column=col).column_letter].width = max_len + 2

        workbook.save(resultados_path)
        
        print(f"Reporte generado: {resultados_path}")

if __name__ == "__main__":
    main()